import pandas as pd
from openpyxl import load_workbook

def mark_attendance(file_path, sheet_name, student_ids, session_date):
    # Load the workbook and the specific sheet
    workbook = load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
        return
    sheet = workbook[sheet_name]

    # Ensure the session date is in MM/DD/YYYY format
    session_date_str = pd.to_datetime(session_date).strftime('%m/%d/%Y')

    # Add a new session date column if it doesn't exist
    if session_date_str not in [cell.value for cell in sheet[1]]:
        # Add the new column header at the end of the first row
        sheet.cell(row=1, column=sheet.max_column + 1, value=session_date_str)

    # Find the column index for the session date
    session_col = [cell.value for cell in sheet[1]].index(session_date_str) + 1

    # Mark attendance for each student ID
    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
        student_id_cell = sheet.cell(row=row, column=1)  # Assuming column 1 contains IDs
        if str(student_id_cell.value) in student_ids:
            sheet.cell(row=row, column=session_col, value='X')  # Mark attendance

    # Save the workbook
    workbook.save(file_path)
    print(f"Attendance marked for session {session_date_str}.")

# Interactive part: Collecting user input
file_path = 'C:/Users/micha/OneDrive/Desktop/study/study/Canada/Memorial University/Career Progression/SI automation/SI Attendence Automation.xlsx'  # Path to your Excel file

# Get the sheet name (course) from the user
sheet_name = input("Enter the course sheet name (e.g., PHYS 3000, ENGI 1050): ")

# Get the session date from the user
session_date = input("Enter the session date (MM/DD/YYYY): ")

# Get student IDs (multiple IDs separated by commas)
student_ids_input = input("Enter student IDs (comma-separated, e.g., 201737715, 202115036): ")

# Convert input into a list of strings (student IDs)
student_ids = [id.strip() for id in student_ids_input.split(',')]

# Call the function to mark attendance
mark_attendance(file_path, sheet_name, student_ids, session_date)
